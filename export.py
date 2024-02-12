import logging
import os
from datetime import datetime
from itertools import groupby
from operator import itemgetter
from os.path import exists

import openpyxl
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.worksheet.worksheet import Worksheet
from scrapy.utils.project import get_project_settings
from tqdm import tqdm

import logging
import os
import re

import fitz
import pandas as pd
from openpyxl.reader.excel import load_workbook
from tqdm import tqdm
import logging
from datetime import datetime
from itertools import groupby
from operator import itemgetter
from os.path import exists

import openpyxl
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.worksheet.worksheet import Worksheet

from RecorderScraper.helpers import load_scraped_data_from_jsonl, parse_recording_date, remove_non_az, format_whitespaces

logger = logging.getLogger(__name__)


def auto_fit_columns(worksheet: Worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(100, length * 1.33)


def set_table_style(worksheet: Worksheet, table_name: str, table_style: str = 'TableStyleMedium2'):
    table_style = TableStyleInfo(name=table_style,
                                 showFirstColumn=False,
                                 showLastColumn=False,
                                 showRowStripes=True,
                                 showColumnStripes=False)
    worksheet.freeze_panes = 'A2'
    table = Table(displayName=table_name, ref=worksheet.dimensions)
    table.tableStyleInfo = table_style
    worksheet.add_table(table)


def prettify_columns(input_names: list[str]):
    return [n.replace('_', ' ').title() for n in input_names]


def is_llc(input_str: str) -> bool:
    words = [remove_non_az(w) for w in input_str.split(' ')]
    to_exclude_list = [
        'TRUSTEE',
        'TRUST',
    ]
    if any([to_exclude in words for to_exclude in to_exclude_list]):
        return False

    entity_list = {
        'LLC',
        'ENTERPRISE',
        'INC',
        'INCORPORATED',
        'CORP',
        'CORPORATION',
        'GROUP',
        'INVEST',
        'INVESTING',
        'SOLUTIONS',
        'LP',
        'FOUNDATION',
        'COMPANY',
        'AMERICAN',
        'ASSOCIATION',
        'DEVELOPMENT',
        'ENTERPRISES',
        'VENTURES',
        'FUND',
        'FUNDING',
        'PROPERTIES',
        'PROPERTY',
        'CAPITAL',
    }
    if any([entity in words for entity in entity_list]):
        return True

    return False


def _generate_excel_report(scraped_data: list[dict[str, str]], scraping_mode: str, output_excel_file_path: str = 'output.xlsx', filter_individuals: bool = False, filter_deed_of_trust_only: bool = False):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook['Sheet'])

    results_sheet_columns = ['keyword', 'county', 'scraping_date', scraping_mode, 'recording_date', 'document_type', 'document_url']

    # Grantee/Grantor Results Sheet
    results_sheet = workbook.create_sheet('Results')
    results_sheet.append(prettify_columns(results_sheet_columns))

    unique_names = set()
    all_results = []
    deed_of_trust_docs = [
        'TRUSTDEED',
        'DEEDOFTRUST',
        'TRUSTDEEDDEEDOFTRUST',
    ]
    for current_record in scraped_data:
        current_names_list = current_record.get(scraping_mode)
        if current_record.get('last_record') or not current_names_list:
            continue

        current_record_docs = [remove_non_az(d) for d in current_record.get('document_type').split(',')]

        if not any(doc_type in deed_of_trust_docs for doc_type in current_record_docs) and filter_deed_of_trust_only:
            continue

        for name in current_names_list:
            name = name.replace('L L C', 'LLC')

            if name in unique_names:
                continue

            unique_names.add(name)

            if not is_llc(name) and filter_individuals:
                continue

            current_result = dict.fromkeys(results_sheet_columns)
            for key in current_result.keys():
                value = current_record.get(key)
                if key == 'recording_date':
                    current_result[key] = parse_recording_date([value])
                elif key == 'scraping_date':
                    current_result[key] = datetime.now().date()
                else:
                    current_result[key] = value

            current_result[scraping_mode] = name

            all_results.append(current_result)

    all_results.sort(key=itemgetter('county', scraping_mode, 'keyword'))
    for result in all_results:
        results_sheet.append(list(result.values()))

    set_table_style(results_sheet, 'results')
    auto_fit_columns(results_sheet)

    # County Counter Sheet
    results_count_by_county_sheet = workbook.create_sheet(f'{scraping_mode.title()} Count by County')
    results_count_by_county_sheet.append(['County', f'Names Found'])
    for key, group in groupby(all_results, key=itemgetter('county')):
        results_count_by_county_sheet.append([key, len(list(group))])

    set_table_style(results_count_by_county_sheet, 'county')
    auto_fit_columns(results_count_by_county_sheet)

    # Lender Counter Sheet
    documents_found_sheet = workbook.create_sheet('Documents Count by Lender')
    documents_found_sheet.append(['Lender(Keyword)', 'Start Date', 'End Date', 'Documents Found'])
    results = [
        list(key) + [len([i for i in group if not i['last_record']])]
        for key, group in groupby(sorted(scraped_data, key=itemgetter('keyword', 'start_date', 'end_date')),
                                  key=itemgetter('keyword', 'start_date', 'end_date'))
    ]
    results.sort(key=lambda x: x[3], reverse=True)
    for result in results:
        documents_found_sheet.append(result)

    set_table_style(documents_found_sheet, 'lender')
    auto_fit_columns(documents_found_sheet)

    workbook.save(output_excel_file_path)


def generate_excel_report(filter_deed_of_trust_only: bool = False, filter_individuals: bool = False, output_excel_file_path: str = ''):
    logger.info('Preparing the Excel report. Please wait...')
    settings = get_project_settings()
    scraping_mode = settings.get('SCRAPING_MODE')
    scraped_data_file_name = settings.get('DATA_FILE')

    if not output_excel_file_path:
        output_excel_file_path = settings.get('OUTPUT_FILE')

    scraped_data = load_scraped_data_from_jsonl(scraped_data_file_name)

    if exists(scraped_data_file_name + 'old'):
        scraped_data += load_scraped_data_from_jsonl(scraped_data_file_name + 'old')

    _generate_excel_report(scraped_data, scraping_mode=scraping_mode, output_excel_file_path=output_excel_file_path, filter_deed_of_trust_only=filter_deed_of_trust_only, filter_individuals=filter_individuals)
    logger.info(f'All data was successfully saved to {output_excel_file_path}')


def _parse_manager_and_members_table(input_file: str):
    doc = fitz.open(input_file)
    data = []
    for page in doc:
        tables = page.find_tables()

        for table in tables:
            current_table_list = table.extract()
            if current_table_list[0][0] == 'Manager or Member Name':
                for index, line in enumerate(current_table_list):
                    if len(line) != 2:
                        raise Exception('TABLE EXCEPTION')
                    if index == 0:
                        continue

                    address_lines = line[1].split('\n')
                    if len(address_lines) != 2 and len(address_lines) != 3:
                        raise Exception('ADDRESS EXCEPTION')

                    name = format_whitespaces(line[0].replace('+', ' ').replace('-', ' '))

                    address1 = format_whitespaces(address_lines[0])
                    address2 = format_whitespaces(address_lines[1]) if len(address_lines) == 3 else ''

                    city_state_zip = format_whitespaces(address_lines[-1])

                    current_line = [os.path.basename(input_file), name, address1, address2, city_state_zip]
                    data.append(current_line)

                return data
    return None
def export_manager_and_member_data(pdf_dir: str, output_file: str, file_matching_dict: dict[str, tuple[str, str]]):
    data = []
    for input_file in tqdm([os.path.join(pdf_dir, input_file) for input_file in os.listdir(pdf_dir)]):
        try:
            filename = os.path.basename(input_file).split('_')[0]
            keyword, county = file_matching_dict[filename]
            current_lines = _parse_manager_and_members_table(input_file)
            if current_lines:
                for _line in current_lines:
                    _line.insert(0, county)
                    _line.insert(0, keyword)
                data.extend(current_lines)

        except Exception as ex:
            logging.exception(ex)

    df = pd.DataFrame(data, columns=['Keyword', 'County', 'PDF File', 'Member Name', 'Address1', 'Address2', 'City, State, Zip'])
    df.to_excel(output_file, index=False)

    existing_workbook = load_workbook(output_file)
    existing_sheet = existing_workbook.active
    auto_fit_columns(existing_sheet)
    set_table_style(existing_sheet, "Table1")
    existing_workbook.save(output_file)


def main():
    generate_excel_report()


if __name__ == '__main__':
    main()
